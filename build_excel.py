#!/usr/bin/env python3
"""
Bygger om data/agartracker.xlsx utifrån data/history_<key>.json för varje
aktie i track_owners.STOCKS, en flik per aktie.

Körs automatiskt efter varje uppdatering i track_owners.py, men kan
också köras fristående:

    python build_excel.py
"""

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment

DATA_DIR = Path(__file__).parent / "data"
EXCEL_FILE = DATA_DIR / "agartracker.xlsx"

FONT_NAME = "Arial"


def _stock_list():
    """Läser STOCKS från track_owners utan att kräva aiohttp/pyavanza."""
    import track_owners
    return track_owners.STOCKS


def build_sheet(wb: Workbook, stock: dict, is_first: bool):
    key = stock["key"]
    name = stock["name"]
    history_path = DATA_DIR / f"history_{key}.json"
    history = json.loads(history_path.read_text(encoding="utf-8")) if history_path.exists() else []

    ws = wb.active if is_first else wb.create_sheet()
    ws.title = name[:31]  # Excel-bladnamn max 31 tecken

    headers = ["Datum", "Avanza", "Nordnet", "Totalt"]
    header_fill = PatternFill(start_color="1F2933", end_color="1F2933", fill_type="solid")
    header_font = Font(name=FONT_NAME, bold=True, color="FFFFFF")

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i, entry in enumerate(history, start=2):
        ws.cell(row=i, column=1, value=entry["date"]).font = Font(name=FONT_NAME)
        ws.cell(row=i, column=2, value=entry["avanza"]).font = Font(name=FONT_NAME)
        ws.cell(row=i, column=3, value=entry["nordnet"]).font = Font(name=FONT_NAME)
        total_cell = ws.cell(row=i, column=4, value=f"=B{i}+C{i}")
        total_cell.font = Font(name=FONT_NAME)

    last_row = len(history) + 1
    for col, width in {"A": 14, "B": 12, "C": 12, "D": 12}.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"

    if last_row >= 2:
        chart = LineChart()
        chart.title = f"{name} – antal ägare över tid"
        chart.style = 12
        chart.y_axis.title = "Antal ägare"
        chart.x_axis.title = "Datum"
        chart.height = 10
        chart.width = 22

        data_ref = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=last_row)
        cats = Reference(ws, min_col=1, min_row=2, max_row=last_row)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)
        chart.series[0].graphicalProperties.line.solidFill = "FF3B30"
        chart.series[1].graphicalProperties.line.solidFill = "34C9EB"

        ws.add_chart(chart, "F2")


def main():
    stocks = _stock_list()
    wb = Workbook()
    for i, stock in enumerate(stocks):
        build_sheet(wb, stock, is_first=(i == 0))

    EXCEL_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(EXCEL_FILE)
    print(f"Excel-fil sparad: {EXCEL_FILE} ({len(stocks)} aktier)")


if __name__ == "__main__":
    main()
