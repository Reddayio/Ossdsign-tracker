#!/usr/bin/env python3
"""
Bygger om data/ossdsign_owners.xlsx utifrån data/history.json.

Körs automatiskt efter varje uppdatering i track_owners.py, men kan
också köras fristående:

    python build_excel.py
"""

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HISTORY_FILE = Path(__file__).parent / "data" / "history.json"
EXCEL_FILE = Path(__file__).parent / "data" / "ossdsign_owners.xlsx"

FONT_NAME = "Arial"


def build_workbook(history: list) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ägardata"

    headers = ["Datum", "Avanza", "Nordnet", "Totalt"]
    header_fill = PatternFill(start_color="1F2933", end_color="1F2933", fill_type="solid")
    header_font = Font(name=FONT_NAME, bold=True, color="FFFFFF")

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i, entry in enumerate(history, start=2):
        ws.cell(row=i, column=1, value=entry["date"]).font = Font(name=FONT_NAME)
        ws.cell(row=i, column=2, value=entry["avanza"]).font = Font(name=FONT_NAME)
        ws.cell(row=i, column=3, value=entry["nordnet"]).font = Font(name=FONT_NAME)
        # Formel, inte hårdkodat värde -> räknas om automatiskt i Excel
        total_cell = ws.cell(row=i, column=4, value=f"=B{i}+C{i}")
        total_cell.font = Font(name=FONT_NAME)

    last_row = len(history) + 1

    # Kolumnbredder
    widths = {"A": 14, "B": 12, "C": 12, "D": 12}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"

    if last_row >= 2:
        chart = LineChart()
        chart.title = "OssDsign – antal ägare över tid"
        chart.style = 12
        chart.y_axis.title = "Antal ägare"
        chart.x_axis.title = "Datum"
        chart.height = 10
        chart.width = 22

        data = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=last_row)
        cats = Reference(ws, min_col=1, min_row=2, max_row=last_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        # Färgsätt serierna: Avanza röd, Nordnet blå
        chart.series[0].graphicalProperties.line.solidFill = "FF3B30"
        chart.series[1].graphicalProperties.line.solidFill = "34C9EB"

        ws.add_chart(chart, f"F2")

    return wb


def main():
    if not HISTORY_FILE.exists():
        history = []
    else:
        history = json.loads(HISTORY_FILE.read_text(encoding="utf-8"))

    wb = build_workbook(history)
    EXCEL_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(EXCEL_FILE)
    print(f"Excel-fil sparad: {EXCEL_FILE} ({len(history)} rader)")


if __name__ == "__main__":
    main()
